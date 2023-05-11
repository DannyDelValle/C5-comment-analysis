import sys
import re
import os
import spacy
import pyreadr
import warnings
import openpyxl
import numpy as np
import pandas as pd
import datetime
from copy import deepcopy
from PyQt5.QtCore import Qt
from autocorrect import Speller
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QFileDialog, QLabel, QComboBox, QProgressBar, QGroupBox
from PyQt5.QtGui import QFont

warnings.filterwarnings('ignore')

nlp = spacy.load("es_core_news_sm")
spell = Speller(lang='es')

def lematizador(textos):
    i=0
    comentarios_lema = []
    for comentario in textos:
        i+=1
        print(i, "/", len(textos), end='\r')
        doc = nlp(comentario)
        comentario_aux = ''
        for token in doc:
            comentario_aux += token.lemma_ + ' '
        comentarios_lema.append(comentario_aux)
    return comentarios_lema

def temporalidad():
    fecha=datetime.date.today()
    dayw = fecha.weekday()
    fecha_lastsunday = fecha-datetime.timedelta(days=dayw)  
    fecha=fecha-datetime.timedelta(days=dayw+1)
    fecha_lastsunday = str(fecha_lastsunday.strftime('%Y-%m-%d'))
    fecha_fin = str(fecha.strftime('%Y-%m-%d')) 
    fecha_inicio = str(fecha.replace(day=1, month=fecha.month-1).strftime('%Y-%m-%d'))
    print(fecha_inicio, fecha_fin)
    return fecha_inicio, fecha_fin, fecha_lastsunday

def porcentaje(valor, total):
        if total == 0:
            return 0
        else:
            return valor/total
        
def graficas_tablas(df_graf, C2, fecha_inicio, fecha_fin, sheet, wb):
    meses = df_graf['Mes'].unique()

    sheet['A2'] = meses[0]
    sheet['A4'] = meses[1]

    sheet['C2'] = len(df_graf[(df_graf['Mes'] == meses[0]) & (df_graf['clas_con_f_alarma2']  == 'DELITOS')])
    sheet['C3'] = len(df_graf[(df_graf['Mes'] == meses[0]) & (df_graf['clas_con_f_alarma2']  == 'OTROS')])
    sheet['C4'] = len(df_graf[(df_graf['Mes'] == meses[1]) & (df_graf['clas_con_f_alarma2']  == 'DELITOS')])
    sheet['C5'] = len(df_graf[(df_graf['Mes'] == meses[1]) & (df_graf['clas_con_f_alarma2']  == 'OTROS')])
    sheet['C6'] = '=SUM(C2:C5)'

    sheet['D2'] = len(df_graf[(df_graf['Mes'] == meses[0]) & (df_graf['codigo_cierre'] == 'A') & (df_graf['clas_con_f_alarma2']  == 'DELITOS')])
    sheet['D3'] = len(df_graf[(df_graf['Mes'] == meses[0]) & (df_graf['codigo_cierre'] == 'A') & (df_graf['clas_con_f_alarma2']  == 'OTROS')])
    sheet['D4'] = len(df_graf[(df_graf['Mes'] == meses[1]) & (df_graf['codigo_cierre'] == 'A') & (df_graf['clas_con_f_alarma2']  == 'DELITOS')])
    sheet['D5'] = len(df_graf[(df_graf['Mes'] == meses[1]) & (df_graf['codigo_cierre'] == 'A') & (df_graf['clas_con_f_alarma2']  == 'OTROS')])
    sheet['D6'] = '=SUM(D2:D5)'

    sheet['E2'] = len(df_graf[(df_graf['Mes'] == meses[0]) & (df_graf['Entrevista'] == 1) & (df_graf['clas_con_f_alarma2'] == 'DELITOS')])
    sheet['E3'] = len(df_graf[(df_graf['Mes'] == meses[0]) & (df_graf['Entrevista'] == 1) & (df_graf['clas_con_f_alarma2'] == 'OTROS')])
    sheet['E4'] = len(df_graf[(df_graf['Mes'] == meses[1]) & (df_graf['Entrevista'] == 1) & (df_graf['clas_con_f_alarma2'] == 'DELITOS')])
    sheet['E5'] = len(df_graf[(df_graf['Mes'] == meses[1]) & (df_graf['Entrevista'] == 1) & (df_graf['clas_con_f_alarma2'] == 'OTROS')])
    sheet['E6'] = '=SUM(E2:E5)'

    sheet['F2'] = '=D2-E2'
    sheet['F3'] = '=D3-E3'
    sheet['F4'] = '=D4-E4'
    sheet['F5'] = '=D5-E5'
    sheet['F6'] = '=D6-E6'

    sheet['G2'] = '=F2/C2'
    sheet['G3'] = '=F3/C3'
    sheet['G4'] = '=F4/C4'
    sheet['G5'] = '=F5/C5'
    sheet['G6'] = '=F6/C6'

    if sheet['G4'].value > sheet['G2'].value:
        sheet['G4'].fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
    else:
        sheet['G4'].fill = PatternFill(start_color='92d050', end_color='92d050', fill_type='solid')

    if sheet['G5'].value > sheet['G3'].value:
        sheet['G5'].fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
    else:
        sheet['G5'].fill = PatternFill(start_color='92d050', end_color='92d050', fill_type='solid')

    sheet['G11'] = len(df_graf[(df_graf['Mes'] == meses[0]) & (df_graf['Amplio recorrido'] == 0)])/(len(df_graf[(df_graf['Mes'] == meses[0])]))
    sheet['G12'] = len(df_graf[(df_graf['Mes'] == meses[1]) & (df_graf['Amplio recorrido'] == 0)])/(len(df_graf[(df_graf['Mes'] == meses[1])]))

    sheet['G15'] = len(df_graf[(df_graf['Mes'] == meses[0]) & (df_graf['Entrevista'] == 1)])/(len(df_graf[(df_graf['Mes'] == meses[0])]))
    sheet['G16'] = len(df_graf[(df_graf['Mes'] == meses[1]) & (df_graf['Entrevista'] == 1)])/(len(df_graf[(df_graf['Mes'] == meses[1])]))

    sheet['G28'] = len(df_graf[df_graf['codigo_cierre'] == 'A'])
    sheet['G29'] = len(df_graf[df_graf['codigo_cierre'] == 'F'])
    sheet['G30'] = len(df_graf[df_graf['codigo_cierre'] == 'D & I']) or (len(df_graf[df_graf['codigo_cierre'] == 'D']) + len(df_graf[df_graf['codigo_cierre'] == 'I']))

    sheet['G34'] = len(df_graf[(df_graf['Amplio recorrido'] == 1)])
    sheet['G35'] = len(df_graf[(df_graf['Amplio recorrido'] == 0)])

    sheet['G38'] = len(df_graf[(df_graf['Categoria'] == 'RE')])
    sheet['G39'] = len(df_graf[(df_graf['Categoria'] == 'RNE')])
    sheet['G40'] = len(df_graf[(df_graf['Categoria'] == 'NRE')])
    sheet['G41'] = len(df_graf[(df_graf['Categoria'] == 'NRNE')])

    wb.save(f'{C2}/NSA_{C2}_{fecha_inicio}_{fecha_fin}.xlsx')

def procesamiento(fecha_inicio, fecha_fin, C2, ruta, rutaCh, processBar, fecha_lastsunday):
    negatives=['(NO|NADIE|NEGATIVO|SIN|AUSENCIA)\s*([SEÉéLQUDHYBRAlI]*\s)*\s*[LOGRACEPNXIM]*\s*[GUNLQIEOA]*\s*[ÉEL]*\s*ENTREVISTA[RSE]*|(NEGATIVO DE QUE ALGUIEN SALGA A ENTREVISTARSE)|(ausencia\s*(de)?\s*alguno\s*(persona)?\s*(con)?\s*(el)?\s*(que)?\s*(él)?\s*entrevista)']
    reasons = ['A\s*M\s*P\s*L\s*I\s*O\s*E\s*L\s*R\s*E\s*C\s*O\s*R\s*R\s*I\s*D\s*O',
            'AMPLI[OARN]+\s*(NUEVAMENTE)?\s*[DNELA\s*]*\s*(RECORRIDO|BUSQUEDA|PERIMETRO|INSPECCION|REVISON|LUGAR|INVESTIGACION|REVISION|PARA DESCARTAR)']
    asi = 'AFIRMATIVO SIN INTERVENCIÓN'
    aci = 'AFIRMATIVO CON INTERVENCIÓN'
    monts = {'01': 'Enero', '02': 'Febrero', '03': 'Marzo', '04': 'Abril', '05': 'Mayo', '06': 'Junio', '07': 'Julio', '08': 'Agosto', '09': 'Septiembre', '10': 'Octubre', '11': 'Noviembre', '12': 'Diciembre'}
    negatives = [re.compile(negative, re.IGNORECASE) for negative in negatives]
    reasons = [re.compile(reson, re.IGNORECASE) for reson in reasons]


    diccionario = pd.read_excel('diccionario.xlsx', dtype=str)
    processBar.setValue(3)
    df = pd.read_csv(ruta, sep=',')

    delegacionesC2 ={'C2 Sur' : 'BENITO JUAREZ, COYOACAN, TLALPAN',
                'C2 Norte' : 'IZTACALCO, GUSTAVO A. MADERO, VENUSTIANO CARRANZA',
                'C2 Centro' : 'CUAUHTEMOC',
                'C2 Oriente' : 'IZTAPALAPA, XOCHIMILCO, MILPA ALTA, TLAHUAC',
                'C2 Poniente' : 'ALVARO OBREGON, AZCAPOTZALCO, MAGDALENA CONTRERAS, MIGUEL HIDALGO, CUAJIMALPA',
                'C2 Centro Histórico' : 'CUAUHTEMOC, VENUSTIANO CARRANZA'
                }

    if not os.path.exists(C2):
        os.makedirs(C2)
    
    if C2 == 'C2 Centro' or C2 == 'C2 Norte' or C2 == 'C2 Centro Histórico':
        result = pyreadr.read_r(rutaCh)
        data = next(iter(result.values()))
        data = data[data['c2_final'] == 'C2 Centro Histórico']
        data = data.reset_index(drop=True)
        if C2 != 'C2 Centro Histórico':
            # eliminar de df los registros que tengan el mismo folio de df y data y tengan en la columna c2_final C2 Centro Histórico
            df = df[~df['folio'].isin(data['folio'])]
            df = df.reset_index(drop=True)
    if C2 == 'C2 Centro Histórico':
        df = df[df['folio'].isin(data['folio'])]

    delegaciones = delegacionesC2[C2]
    delegaciones = delegaciones.split(', ')
    df = df[df['delegacion_cierre'].isin(delegaciones)]

    df['Mes'] = df['fecha_cierre'].astype(str).str.extract(r'-(\d\d)', expand=False)
    df['Mes'] = df['Mes'].map(monts)
    df = df[(df['fecha_cierre'] >= fecha_inicio) & (df['fecha_cierre'] <= fecha_lastsunday)]

    df_copy = df.copy()

    # A = len(df[df['codigo_cierre'] == 'A'])
    # D = len(df[df['codigo_cierre'] == 'D'])
    # I = len(df[df['codigo_cierre'] == 'I'])
    # F = len(df[df['codigo_cierre'] == 'F'])

    df = df[df['codigo_cierre'] == 'A']
    df = df.reset_index(drop=True)
    df_copy = df_copy.reset_index(drop=True)
    # print("A: ",A, "\tD: ",D, "\tI: ",I, "\tF: ",F)
    # actualizar barra de progreso 
    processBar.setValue(10)

    df = df.replace('Ã\x81', 'Á', regex=True)
    df = df.replace('Ã¡', 'á', regex=True)
    df = df.replace('Ã‰', 'É', regex=True)
    df = df.replace('Ã©', 'é', regex=True)
    df = df.replace('Ã\x8d', 'Í', regex=True)
    df = df.replace('Ã\xad', 'í', regex=True)
    df = df.replace('Ã“', 'Ó', regex=True)
    df = df.replace('Ã³', 'ó', regex=True)
    df = df.replace('Ãš', 'Ú', regex=True)
    df = df.replace('Ãº', 'ú', regex=True)
    df = df.replace('Ã‘', 'Ñ', regex=True)
    df = df.replace('Ã±', 'ñ', regex=True)
    df = df.replace('Â°', '°', regex=True)
    df = df.reset_index(drop=True)


    df_tabla = df[['folio']]
    df_tabla['CR'] = df[['codigo_cierre']]
    df_tabla['CI'] = 0
    df_tabla['CG'] = 0
    df_tabla['Mes'] = df['Mes']
    df_tabla['Clasificación'] = df['clas_con_f_alarma']
    df_tabla.loc[df_tabla['Clasificación'].str.contains('DELITO', na=False), 'Clasificación'] = 'DELITOS'
    df_tabla.loc[df_tabla['Clasificación'] != 'DELITOS', 'Clasificación'] = 'OTROS'
    df_tabla['Amplio recorrido'] = np.nan
    df_tabla['Entrevista'] = np.nan
    df_tabla['Incidente'] = df[['incidente_c4']]

    df_tabla['CG'] = df['comentarios.y'].str.extract(r'CIERRE GLOBAL: (.+?);', expand=False)

    df_tabla.loc[df_tabla['CG'].str.contains(asi, na=False), 'CI'] = 'ASI'
    df_tabla.loc[df_tabla['CG'].str.contains(aci, na=False), 'CI'] = 'ACI'
    df_tabla.loc[df_tabla['CI'] == 0, 'CI'] = np.nan

    elementos = ['SE AGREGÓ UNA DESCRIPCIÓN: ', 'SE AGREGÓ NOTA DE CIERRE: ', 'SE AGREGÓ CIERRE GLOBAL: ',]
    for i in range(len(df)):
        # df['comentarios.y'][i] = df['comentarios.y'][i][df['comentarios.y'][i].find('NADIE SOLICITA EL APOYO'):]
        df['comentarios.y'][i] = re.split('\[\d{2}\/\d{2}\/\d{4}\s\d{2}:\d{2}:\d{2}\]\s', df['comentarios.y'][i])
        df_aux = []
        for j in range(len(df['comentarios.y'][i])):
            if any(elemento in df['comentarios.y'][i][j] for elemento in elementos):
                df_aux.append(df['comentarios.y'][i][j])
        df['comentarios.y'][i] = df_aux
        df['comentarios.y'][i] = [x for x in df['comentarios.y'][i] if any(elemento in x for elemento in elementos)]
    processBar.setValue(20)
    
    for comentario in range(len(df)):
        for j in range(len(df['comentarios.y'][comentario])):
            df['comentarios.y'][comentario][j] = re.sub('\[\w+\](\s*[\wÀ-Ý]+\s*)+:\s*', '', df['comentarios.y'][comentario][j])
        df['comentarios.y'][comentario] = ' '.join(df['comentarios.y'][comentario])
        df['comentarios.y'][comentario] = re.sub('[^a-zA-ZñÑáéíóúÁÉÍÓÚ\s]', '', df['comentarios.y'][comentario])
        df['comentarios.y'][comentario] = re.sub('\s+', ' ', df['comentarios.y'][comentario])
    df['comentarios.y'] = df['comentarios.y'].str.lower()
    processBar.setValue(30)

    df['comentarios.y'] = lematizador(df['comentarios.y'])

    for i in range(len(df)):
        df['comentarios.y'][i] = re.sub('nombre\w+', 'nombre', df['comentarios.y'][i])
        df['comentarios.y'][i] = re.sub('\w+tipo', 'tipo', df['comentarios.y'][i])
        df['comentarios.y'][i] = re.sub('submarca\w+', 'submarca', df['comentarios.y'][i])
        df['comentarios.y'][i] = re.sub('marca\w+', 'marca', df['comentarios.y'][i])
        df['comentarios.y'][i] = re.sub('placa\w+', 'placa', df['comentarios.y'][i])
        df['comentarios.y'][i] = re.sub('\w+masculino', 'masculino', df['comentarios.y'][i])
        df['comentarios.y'][i] = re.sub('\w+femenino', 'femenino', df['comentarios.y'][i])
        df['comentarios.y'][i] = re.sub('color\w+', 'color', df['comentarios.y'][i])
        df['comentarios.y'][i] = re.sub('\w+color', 'color', df['comentarios.y'][i])
        df['comentarios.y'][i] = re.sub('colonia\w+', 'colomia', df['comentarios.y'][i])
        df['comentarios.y'][i] = re.sub('particulares\w+', 'particulares', df['comentarios.y'][i])
        df['comentarios.y'][i] = re.sub('http[\w./\-#;:%_&$!?¡¿"()[]]+', '', df['comentarios.y'][i])
        df['comentarios.y'][i] = re.sub('\w+informar', 'informar', df['comentarios.y'][i])

    palabras_no_encontradas = []
    for i in range(len(df)):
        print(i+1, "/", len(df), end="\r")
        for palabra in df['comentarios.y'][i].split():
            # si la palabra está en el diccionario, reemplazarla por la corrección solo si la plabra esta entre espacios
            if palabra in diccionario['palabra'].values:
                df['comentarios.y'][i] = re.sub(r'\b' + palabra + r'\b', str(diccionario['corrección'][diccionario['palabra'] == palabra].values[0]), df['comentarios.y'][i])
            else:
                if palabra not in palabras_no_encontradas:
                    palabras_no_encontradas.append(palabra)
                    aux = palabra
                    palabra = spell(palabra)
                    if palabra in diccionario['palabra'].values:
                        df['comentarios.y'][i] = re.sub(r'\b' + palabra + r'\b', str(diccionario['corrección'][diccionario['palabra'] == palabra].values[0]), df['comentarios.y'][i])
                    else:
                        doc = nlp(palabra)
                        for token in doc:
                            palabra = token.lemma_
                        if palabra in diccionario['palabra'].values:
                            df['comentarios.y'][i] = re.sub(r'\b' + palabra + r'\b', str(diccionario['corrección'][diccionario['palabra'] == palabra].values[0]), df['comentarios.y'][i])
                            palabras_no_encontradas.remove(aux)
    processBar.setValue(80)
    df['Amplio recorrido'] = np.nan
    df['Entrevista'] = np.nan
    df['Categoria'] = np.nan

    df_copy['Amplio recorrido'] = np.nan
    df_copy['Entrevista'] = np.nan
    df_copy['Categoria'] = np.nan

    matches = []
    matches2 = []
    matches3 = []

    for i in range(len(df)):
        for reason in reasons:
            if reason.search(df['comentarios.y'][i]):
                # print(i)
                # matches.append(df_copy['comentarios.y'][i])
                matches.append(reason.search(df_copy['comentarios.y'][i]))
                df_tabla['Amplio recorrido'][i] = 1
                df['Amplio recorrido'][i] = 1
                break

    for i in range(len(df)):
        for negative in negatives:
            if negative.search(df['comentarios.y'][i]):
                df_tabla['Entrevista'][i] = 0
                # matches2.append(df_copy['comentarios.y'][i])
                matches.append(negative.search(df_copy['comentarios.y'][i]))
                break
        if 'entrevista' in df['comentarios.y'][i] and df_tabla['Entrevista'][i] != 0:
            # print(i)
            matches3.append(df['comentarios.y'][i][df['comentarios.y'][i].find('entrevista')-40:df['comentarios.y'][i].find('entrevista')+40])
            df_tabla['Entrevista'][i] = 1
            df['Entrevista'][i] = 1

    # fill nan values with 0
    df_tabla['Entrevista'] = df_tabla['Entrevista'].fillna(0)
    df_tabla['Amplio recorrido'] = df_tabla['Amplio recorrido'].fillna(0)

    df['Amplio recorrido'] = df['Amplio recorrido'].fillna(0)
    df['Entrevista'] = df['Entrevista'].fillna(0)

    # categorias RE, RNE, NRE, NRNE
    for i in range(len(df)):
        if df['Amplio recorrido'][i] == 1 and df['Entrevista'][i] == 1:
            df['Categoria'][i] = 'RE'
        elif df['Amplio recorrido'][i] == 1 and df['Entrevista'][i] == 0:
            df['Categoria'][i] = 'RNE'
        elif df['Amplio recorrido'][i] == 0 and df['Entrevista'][i] == 1:
            df['Categoria'][i] = 'NRE'
        elif df['Amplio recorrido'][i] == 0 and df['Entrevista'][i] == 0:
            df['Categoria'][i] = 'NRNE'

    for i in range(len(df)):
        for j in range(len(df_copy)):
            if df['folio'][i] == df_copy['folio'][j]:
                df_copy['Amplio recorrido'][j] = df['Amplio recorrido'][i]
                df_copy['Entrevista'][j] = df['Entrevista'][i]
                df_copy['Categoria'][j] = df['Categoria'][i]
    processBar.setValue(90)

    # cuantos incidentes tienen CR "A"
    # print("Afirmativos: ",df_tabla[df_tabla['CR'] == 'A'].shape[0])
    # print("Amplio recorrido: ",df_tabla[(df_tabla['Amplio recorrido'] == 1) & (df_tabla['CR'] == 'A')].shape[0])
    # print("Entrevista: ",df_tabla[(df_tabla['Entrevista'] == 1) & (df_tabla['CR'] == 'A')].shape[0])

    # # cuantos tienen entrevista y amplio recorrido y CR "A"
    # print("Entrevista y amplio recorrido: ",RE:=df_tabla[(df_tabla['Entrevista'] == 1) & (df_tabla['Amplio recorrido'] == 1) & (df_tabla['CR'] == 'A')].shape[0])
    # # cuantos tienen 0 en entrevista y 1 en amplio recorrido y CR "A"
    # print("No entrevista y amplio recorrido: ",RNE:=df_tabla[(df_tabla['Entrevista'] == 0) & (df_tabla['Amplio recorrido'] == 1) & (df_tabla['CR'] == 'A')].shape[0])
    # # cuantos tienen 1 en entrevista y 0 en amplio recorrido y CR "A"
    # print("Entrevista y no amplio recorrido: ",NRE:=df_tabla[(df_tabla['Entrevista'] == 1) & (df_tabla['Amplio recorrido'] == 0) & (df_tabla['CR'] == 'A')].shape[0])
    # # cuantos tienen 0 en entrevista y 0 en amplio recorrido y CR "A"
    # print("No entrevista y no amplio recorrido: ",NRNE:=df_tabla[(df_tabla['Entrevista'] == 0) & (df_tabla['Amplio recorrido'] == 0) & (df_tabla['CR'] == 'A')].shape[0])
    
    wb = openpyxl.load_workbook('Tabla y graficas.xlsx')
    wb.save(f'{C2}/NSA_{C2}_{fecha_inicio}_{fecha_fin}.xlsx')
    wb = openpyxl.load_workbook(f'{C2}/NSA_{C2}_{fecha_inicio}_{fecha_fin}.xlsx')

    nombres = ['General']

    nombres = nombres + delegaciones

    for nombre in nombres:
        processBar.setValue(90+int(len(nombres)/len(nombres)*10))   
        print(nombre)
        wb.copy_worksheet(wb['Tabla y graficas'])
        wb['Tabla y graficas Copy'].title = nombre
        sheet = wb[nombre]

        #
        for chart in wb['Tabla y graficas']._charts:
            chart_copia = deepcopy(chart)
            # chart_copia.add_data(data, from_rows=True, titles_from_data=True)

            for serie in chart_copia.ser:
                serie.tx.strRef.f = serie.tx.strRef.f.replace('Tabla y graficas', nombre)
                serie.cat.strRef.f = serie.cat.strRef.f.replace('Tabla y graficas', nombre)
                serie.val.numRef.f = serie.val.numRef.f.replace('Tabla y graficas', nombre)

            sheet.add_chart(chart_copia)
        # 

        if nombre == 'General':
            graficas_tablas( df_copy, C2, fecha_inicio, fecha_fin, sheet, wb)
        else:
            df_graf = df_copy[df_copy['delegacion_cierre'] == nombre]
            graficas_tablas( df_graf, C2, fecha_inicio, fecha_fin, sheet, wb)

    wb.create_sheet('Tabla')
    sheet = wb['Tabla']

    for r in dataframe_to_rows(df_tabla, index=False, header=True):
        sheet.append(r)

    wb.create_sheet('NSA')
    sheet2 = wb['NSA']

    for r in dataframe_to_rows(df_copy, index=False, header=True):
        sheet2.append(r)

    del wb['Tabla y graficas']
        
    wb.save(f'{C2}/NSA_{C2}_{fecha_inicio}_{fecha_fin}.xlsx')
    wb.close()
    processBar.setValue(100)    

class VentanaPrincipal(QWidget):
    def __init__(self):
        self.fecha_inicio, self.fecha_fin, self.fecha_lastsunday = temporalidad()
        super().__init__()
        self.initUI()
        self.rutaCH = ''

    def initUI(self):
        self.setWindowTitle('Nadie Solicita el Apoyo')
        self.setGeometry(500, 500, 800, 400)

        layout = QVBoxLayout()
        temporalidad = QLabel('Peridodo de análisis: '+self.fecha_inicio+' a '+self.fecha_fin)
        temporalidad.setFont(QFont('Arial', 20, QFont.Bold))
        temporalidad.setStyleSheet('color: #0099ff')
        layout.addWidget(temporalidad)

        # Botón para seleccionar archivo
        titleBA=QLabel('\nSelecciona el archivo de datos')
        titleBA.setFont(QFont('Arial', 15, QFont.Bold))
        titleBA.setStyleSheet('color: #0d3047')
        layout.addWidget(titleBA)
        self.botonArchivo = QPushButton('Seleccionar archivo', self, clicked=self.abrirArchivo)
        layout.addWidget(self.botonArchivo)
        self.labelArchivo = QLabel('Archivo no seleccionado')
        layout.addWidget(self.labelArchivo)

        # ComboBox para seleccionar C2
        titleC2=QLabel('\nSelecciona C2 a analizar')
        titleC2.setFont(QFont('Arial', 15, QFont.Bold))
        titleC2.setStyleSheet('color: #0d3047')
        layout.addWidget(titleC2)
        self.comboBoxC2 = QLabel('Selecciona C2')
        self.comboBoxC2 = QComboBox(self)
        self.comboBoxC2.addItems(['C2 Poniente', 'C2 Norte', 'C2 Sur', 'C2 Centro', 'C2 Oriente', 'C2 Centro Histórico'])
        self.comboBoxC2.currentIndexChanged.connect(self.seleccionC2)
        layout.addWidget(self.comboBoxC2)

        # Botón para seleccionar archivo RDS
        titleRDS=QLabel('\nSelecciona el archivo RDS')
        titleRDS.setFont(QFont('Arial', 15, QFont.Bold))
        titleRDS.setStyleSheet('color: #0d3047')
        layout.addWidget(titleRDS)
        slabelRDS = QLabel('Solo para C2 Norte, C2 Centro y  C2 Centro Histórico')
        slabelRDS.setFont(QFont('Arial', 8))
        slabelRDS.setStyleSheet('color:#666666')
        layout.addWidget(slabelRDS)
        self.botonArchivoRDS = QPushButton('Seleccionar archivo RDS', self)
        self.botonArchivoRDS.clicked.connect(self.abrirArchivoRDS)
        self.botonArchivoRDS.setEnabled(False)
        layout.addWidget(self.botonArchivoRDS)

        self.labelArchivoRDS = QLabel('Archivo RDS no seleccionado\n\n')
        self.labelArchivoRDS.setEnabled(False)
        layout.addWidget(self.labelArchivoRDS)

        # Botón de aceptar
        self.botonAceptar = QPushButton('Aceptar', self)
        self.botonAceptar.clicked.connect(self.aceptar)
        self.botonAceptar.setStyleSheet('background-color: #00bf29; color: white')
        font = self.botonAceptar.font()
        font.setPointSize(15)
        self.botonAceptar.setFont(font)
        layout.addWidget(self.botonAceptar)

        # Barra de progreso
        self.progressBar = QProgressBar(self)
        layout.addWidget(self.progressBar)

        self.setLayout(layout)

    def abrirArchivo(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        archivo, _ = QFileDialog.getOpenFileName(self, 'Seleccionar archivo', '', 'Todos los archivos (*)', options=options)

        if archivo:
            self.rutaData = archivo
            self.labelArchivo.setText(f'Archivo seleccionado: {archivo}')

    def seleccionC2(self, index):
        self.C2 = self.comboBoxC2.itemText(index)
        if self.C2 in ['C2 Norte', 'C2 Centro', 'C2 Centro Histórico']:
            self.botonArchivoRDS.setEnabled(True)
            self.labelArchivoRDS.setEnabled(True)
        else:
            self.botonArchivoRDS.setEnabled(False)
            self.labelArchivoRDS.setEnabled(False)

    def abrirArchivoRDS(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        archivo, _ = QFileDialog.getOpenFileName(self, 'Seleccionar archivo RDS', '', 'Archivo RDS (*.rds)', options=options)

        if archivo:
            self.rutaCH = archivo
            self.labelArchivoRDS.setText(f'Archivo RDS seleccionado: {archivo}\n\n')


    def aceptar(self):
        self.progressBar.setValue(0)
        procesamiento(self.fecha_inicio, self.fecha_fin, self.C2, self.rutaData, self.rutaCH, self.progressBar, self.fecha_lastsunday)
        self.labelFinalizado = QLabel('Proceso finalizado')
        self.layout().addWidget(self.labelFinalizado)

def main():
    app = QApplication(sys.argv)
    ventana = VentanaPrincipal()
    ventana.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()
